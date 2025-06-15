import os
import time
import requests
import shutil
import json
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from datetime import datetime

API_BASE_URL = 'https://tools.clc.hcmus.edu.vn'
PDF_FILE_NAME = 'VN_HánVăn_TiểuThuyết_TậpThành - 04.pdf'

class HanNomPDFProcessor:
    def __init__(self):
        self.domain = 'L'
        self.subDomain = 'N'
        self.genre = 'H'
        self.fileNumber = '004'
        os.makedirs("temp", exist_ok=True)
        os.makedirs("output", exist_ok=True)
        
        # Tracking variables
        self.processing_log = []
        self.successful_files = []
        self.failed_files = []

    def generate_id(self, chapter: int, page: int, sentence: int) -> str:
        return f"{self.domain}{self.subDomain}{self.genre}_{self.fileNumber}.{chapter:03}.{page:03}.{sentence:02}"

    def retry_with_backoff(self, func, retries=3, base_delay=1.0):
        for attempt in range(1, retries + 1):
            try:
                return func()
            except Exception as e:
                if attempt == retries:
                    raise
                delay = base_delay * (2 ** (attempt - 1))
                print(f"Thử lại lần {attempt + 1} sau {delay:.1f}s...")
                time.sleep(delay)

    # def convert_pdf_to_images(self, pdf_path):
    #     print("Đang chuyển đổi PDF thành hình ảnh...")
    #     images = convert_from_path(pdf_path, dpi=300)
    #     image_paths = []
    #     for i, img in enumerate(images):
    #         path = f"./temp/page{i + 1}.png"
    #         img.save(path, 'PNG')
    #         image_paths.append(path)
    #     return image_paths

    def upload_image(self, image_path):
        def _upload():
            print(f"Uploading: {image_path}")
            with open(image_path, 'rb') as f:
                files = {'image_file': f}
                headers = {'User-Agent': 'HanNom-OCR-Client'}
                response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/image-upload", files=files, headers=headers)
                res_json = response.json()
                if res_json.get("is_success"):
                    uploaded_filename = res_json["data"]["file_name"]
                    print(f"Upload thành công: {uploaded_filename}")
                    return uploaded_filename
                else:
                    raise Exception(f"Upload failed: {res_json.get('message')}")
        return self.retry_with_backoff(_upload, retries=3, base_delay=2)

    def classify_image(self, file_name):
        def _classify():
            payload = {"file_name": file_name}
            headers = {'User-Agent': 'HanNom-OCR-Client', 'Content-Type': 'application/json'}
            response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/image-classification", json=payload, headers=headers)
            res_json = response.json()
            if res_json.get("is_success"):
                print(f"Classification thành công: {file_name}")
                return res_json["data"]
            else:
                raise Exception(f"Classification failed: {res_json.get('message')}")
        return self.retry_with_backoff(_classify, retries=3, base_delay=1)
    
    def preprocessing_image(self, old_file_name):
        def _classify():
            payload = {"file_name": old_file_name}
            headers = {'User-Agent': 'HanNom-OCR-Client', 'Content-Type': 'application/json'}
            response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/image-preprocessing", json=payload, headers=headers)
            res_json = response.json()
            if res_json.get("is_success"):
                print(f"Preprocessing image thành công: {old_file_name}")
                return res_json["data"]["new_file_name"]
            else:
                raise Exception(f"Preprocessing failed: {res_json.get('message')}")
        return self.retry_with_backoff(_classify, retries=3, base_delay=1)

    def ocr_image(self, file_name, ocr_id):
        def _ocr():
            payload = {"ocr_id": ocr_id, "file_name": file_name}
            headers = {'User-Agent': 'HanNom-OCR-Client', 'Content-Type': 'application/json'}
            response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/image-ocr", json=payload, headers=headers)
            res_json = response.json()
            if res_json.get("is_success"):
                return res_json["data"]
            else:
                raise Exception(f"OCR failed: {res_json.get('message')}")
        return self.retry_with_backoff(_ocr, retries=3, base_delay=2)

    # Chuyển văn bản chữ Hán-Nôm thành âm Hán Việt (phiên âm)
    def transliterate(self, text):
        def _translit():
            headers = {'User-Agent': 'HanNom-OCR-Client', 'Content-Type': 'application/json'}
            response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/sinonom-transliteration", json={"text": text}, headers=headers)
            res_json = response.json()
            if res_json.get("is_success"):
                return ' '.join(res_json["data"]["result_text_transcription"])
            else:
                raise Exception(f"Transliterate failed: {res_json.get('message')}")
        
        return self.retry_with_backoff(_translit, retries=3, base_delay=1)

    # Văn xuôi là thể loại văn bản không có vần điệu
    def translate_prose(self, text):
        headers = {'User-Agent': 'HanNom-OCR-Client', 'Content-Type': 'application/json'}
        response = requests.post(f"{API_BASE_URL}/api/web/clc-sinonom/sinonom-prose-translation", json={"text": text, "lang_type": 0}, headers=headers)
        res_json = response.json()
        if res_json.get("is_success"):
            return ' '.join(res_json["data"]["result"])
        else:
            raise Exception(f"Translate prose failed: {res_json.get('message')}")


    def log_processing_status(self, image_path, page_number, status, uploaded_filename=None, error_message=None, processed_count=0):
        """Ghi log trạng thái xử lý của từng file"""
        log_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "image_path": image_path,
            "page_number": page_number,
            "status": status,  # "SUCCESS" hoặc "FAILED"
            "uploaded_filename": uploaded_filename,
            "error_message": error_message,
            "processed_sentences": processed_count
        }
        self.processing_log.append(log_entry)
        
        if status == "SUCCESS":
            self.successful_files.append(log_entry)
        else:
            self.failed_files.append(log_entry)

    def process_page(self, image_path, page_number):
        print(f"Đang xử lý trang {page_number}...")
        results = []
        uploaded_filename = None
        
        try:
            # Upload image và lưu tên file
            uploaded_filename = self.upload_image(image_path)
            
            # Classify và OCR
            classification = self.classify_image(uploaded_filename)
            ocrId = classification["ocr_id"]
            # Preprocessing image
            preprocessed_file_name = self.preprocessing_image(uploaded_filename)
            # OCR
            ocr_data = self.ocr_image(preprocessed_file_name, ocrId)


            if(ocrId == 1 or ocrId == 4): # Thông thường dọc: 1, ngang: 4
                for i, text in enumerate(ocr_data["result_ocr_text"]):
                    if text.strip():
                        print(f"Đoạn {i+1}: {text}")
                        id_ = self.generate_id(1, page_number, i + 1)
                        phonetic = self.transliterate(text)
                        time.sleep(0.5)
                        meaning = self.translate_prose(text)
                        time.sleep(0.5)
                    
                        results.append({
                            "id": id_,
                            "image_box": json.dumps(ocr_data["result_bbox"][i][0]),
                            "sino_nom_ocr": text,
                            "phonetic": phonetic,
                            "meaning": meaning,
                            "uploaded_filename": uploaded_filename,
                            "ocrId": ocrId,
                            "result_ocr_vi_text": ""
                        })
            else:
                # Ngoại cảnh: Không có bounding box
                    # Xử lý từng đoạn text
                for i, text in enumerate(ocr_data["result_ocr_text"]):
                    if text.strip():
                        print(f"Đoạn {i+1}: {text}")
                        id_ = self.generate_id(1, page_number, i + 1)
                        phonetic = self.transliterate(text)
                        time.sleep(0.5)
                        meaning = self.translate_prose(text)
                        time.sleep(0.5)
                    
                        results.append({
                            "id": id_,
                            "image_box": "",
                            "sino_nom_ocr": text,
                            "phonetic": phonetic,
                            "meaning": meaning,
                            "uploaded_filename": uploaded_filename,
                            "ocrId": ocrId,
                            "result_ocr_vi_text": ocr_data["result_ocr_vi_text"][i]
                        })
            
            
            # Log thành công
            self.log_processing_status(
                image_path=image_path,
                page_number=page_number,
                status="SUCCESS",
                uploaded_filename=uploaded_filename,
                processed_count=len(results)
            )
            print(f"✅ Xử lý thành công trang {page_number}: {len(results)} đoạn văn")
            
        except Exception as e:
            error_msg = str(e)
            print(f"❌ Lỗi trang {page_number}: {error_msg}")
            
            # Log thất bại
            self.log_processing_status(
                image_path=image_path,
                page_number=page_number,
                status="FAILED",
                uploaded_filename=uploaded_filename,
                error_message=error_msg,
                processed_count=0
            )
            
        return results

    def export_to_excel(self, results, output_path):
        """Export kết quả OCR ra Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "HanNom OCR Results"
        
        # Headers cho sheet chính
        headers = ["ID", "Image Box", "SinoNom OCR", "Âm Hán Việt", "Nghĩa thuần việt", "Uploaded Filename", "OCR ID", "OCR Vi Text(For 2 only)"]
        ws.append(headers)
        
        # Thêm dữ liệu
        for item in results:
            ws.append([
                item["id"],
                item["image_box"],
                item["sino_nom_ocr"],
                item["phonetic"],
                item["meaning"],
                item.get("uploaded_filename", ""),
                item["ocrId"],
                item["result_ocr_vi_text"]
            ])
        
        wb.save(output_path)
        print(f"✅ Đã lưu kết quả OCR: {output_path}")

    def export_processing_log(self, output_path):
        """Export log xử lý ra Excel"""
        wb = Workbook()
        
        # Sheet 1: Tổng quan
        ws_summary = wb.active
        ws_summary.title = "Tổng quan"
        
        # Thống kê tổng quan
        total_files = len(self.processing_log)
        successful_count = len(self.successful_files)
        failed_count = len(self.failed_files)
        
        summary_data = [
            ["Thống kê xử lý file", ""],
            ["Tổng số file", total_files],
            ["Thành công", successful_count],
            ["Thất bại", failed_count],
            ["Tỷ lệ thành công", f"{(successful_count/total_files*100):.1f}%" if total_files > 0 else "0%"],
            ["", ""],
            ["Thời gian tạo báo cáo", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Định dạng
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Sheet 2: Chi tiết tất cả file
        ws_detail = wb.create_sheet("Chi tiết xử lý")
        detail_headers = ["Thời gian", "Đường dẫn file", "Trang", "Trạng thái", "Tên file uploaded", "Số câu xử lý", "Lỗi"]
        ws_detail.append(detail_headers)
        
        # Màu sắc cho trạng thái
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for i, log in enumerate(self.processing_log, start=2):
            ws_detail.append([
                log["timestamp"],
                log["image_path"],
                log["page_number"],
                log["status"],
                log.get("uploaded_filename", ""),
                log["processed_sentences"],
                log.get("error_message", "")
            ])
            
            # Tô màu theo trạng thái
            status_cell = ws_detail.cell(row=i, column=4)
            if log["status"] == "SUCCESS":
                status_cell.fill = green_fill
            else:
                status_cell.fill = red_fill
        
        # Định dạng cột
        detail_col_widths = [20, 30, 8, 12, 25, 12, 50]
        for i, width in enumerate(detail_col_widths, start=1):
            ws_detail.column_dimensions[get_column_letter(i)].width = width
        
        # Sheet 3: Chỉ file thất bại
        if self.failed_files:
            ws_failed = wb.create_sheet("File thất bại")
            ws_failed.append(detail_headers)
            
            for i, log in enumerate(self.failed_files, start=2):
                ws_failed.append([
                    log["timestamp"],
                    log["image_path"],
                    log["page_number"],
                    log["status"],
                    log.get("uploaded_filename", ""),
                    log["processed_sentences"],
                    log.get("error_message", "")
                ])
                
                # Tô màu đỏ
                status_cell = ws_failed.cell(row=i, column=4)
                status_cell.fill = red_fill
            
            # Định dạng cột
            for i, width in enumerate(detail_col_widths, start=1):
                ws_failed.column_dimensions[get_column_letter(i)].width = width
        
        wb.save(output_path)
        print(f"✅ Đã lưu log xử lý: {output_path}")

    def print_summary(self):
        """In tóm tắt kết quả xử lý"""
        total = len(self.processing_log)
        success = len(self.successful_files)
        failed = len(self.failed_files)
        
        print("\n" + "="*60)
        print("📊 TÓM TẮT XỬ LÝ")
        print("="*60)
        print(f"📁 Tổng số file xử lý: {total}")
        print(f"✅ Thành công: {success}")
        print(f"❌ Thất bại: {failed}")
        if total > 0:
            print(f"📈 Tỷ lệ thành công: {success/total*100:.1f}%")
        print("="*60)
        
        if failed > 0:
            print("\n❌ Danh sách file thất bại:")
            for fail in self.failed_files:
                print(f"  - Trang {fail['page_number']}: {fail['error_message']}")

    def process_pdf(self):
        # Sử dụng danh sách file có sẵn thay vì convert PDF
        image_paths = ["./resized/page4.png"]  # Có thể mở rộng danh sách này
        
        print(f"🚀 Bắt đầu xử lý {len(image_paths)} file...")
        all_results = []
        
        for i, path in enumerate(image_paths):
            results = self.process_page(path, i + 1)
            all_results.extend(results)
            time.sleep(1.5)  # Delay giữa các request
        
        # Export kết quả OCR
        if all_results:
            self.export_to_excel(all_results, "./output/hannom_ocr_results.xlsx")
        
        # Export log xử lý
        self.export_processing_log(f"./output/processing_log{datetime.now()}.xlsx")
        
        # In tóm tắt
        self.print_summary()


if __name__ == "__main__":
    processor = HanNomPDFProcessor()
    try:
        processor.process_pdf()
    except Exception as e:
        print(f"❌ Lỗi chính: {e}")
        # Vẫn export log ngay cả khi có lỗi
        try:
            processor.export_processing_log(f"./output/processing_log{datetime.now()}.xlsx")
        except:
            pass


